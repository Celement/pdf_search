import sys
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import qdarkstyle
import os
import pdfplumber
import re
import pandas as pd
from PyPDF2 import PdfFileReader
import fitz
import openpyxl
from cz import *
class SearchPDFWorker(QThread):
    update_text = pyqtSignal(str)
    update_text_browser = pyqtSignal(str)

    def __init__(self, excel_name, pdf_name):
        super().__init__()
        self.excel_name = excel_name
        self.pdf_name = pdf_name

    def run(self):
        self.search_pdf(self.excel_name, self.pdf_name)

    def get_num_pages(self, file_path):
        reader = PdfFileReader(file_path)
        page_num = reader.getNumPages()
        return page_num

    def extract_text_info(self, filepath, pages):
        with pdfplumber.open(filepath) as pdf:
            page = pdf.pages[pages]
            text = page.extract_text()
            return text

    def pdf_2_image(self, i, image_path, pdf_name):
        PDFdoc = fitz.open(pdf_name)
        page = PDFdoc[i]
        zoom_x = 3
        zoom_y = 3
        mat = fitz.Matrix(zoom_x, zoom_y)
        pix = page.get_pixmap(matrix=mat)
        if not os.path.exists(image_path):
            os.makedirs(image_path)
        pix.save(image_path + "第" + str(i + 1) + "页.png")

    def write_txt(self, file_path, text):
        with open(file_path, 'a', encoding='utf-8') as file:
            file.write(text + '\n')

    def match_pattern(self, pattern, text, i, pdf_name):
        matches = re.findall(pattern, text)
        if matches:
            self.pdf_2_image(i, "./img/", pdf_name)
            self.write_txt("./img/" + "第" + str(i + 1) + "页" + ".txt", pattern)
            return 1
        else:
            return 0

    def check_nan(self, value):
        return value != "nan"

    def search_pdf(self, excel_name, pdf_name):
        flag2list = []
        count = self.get_num_pages(pdf_name)
        for i in range(count):
            text = self.extract_text_info(pdf_name, i)
            df = pd.read_excel(excel_name)
            cont_text = text.strip()
            for index, row in df.iterrows():
                pa = str(row[2]).strip()
                if self.check_nan(pa):
                    flag2 = self.match_pattern(pa, cont_text, i, pdf_name)
                    if flag2 == 1:
                        self.update_text.emit("第" + str(i + 1) + "页找到了" + str(pa) + "\n")
                        flag2list.append(flag2)
                        QThread.msleep(100)  # 添加短暂的延时以确保 UI 有时间更新
                    else:
                        self.update_text_browser.emit("第" + str(i + 1) + "页没有:   " + str(pa) + "\n")

        os.remove(excel_name)
        return flag2list

class QmyWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

    @pyqtSlot()
    def on_pushButton_clicked(self):
        text = self.getText()
        if text == "pdf772s":
            excel_name = self.ui.lineEdit.text()
            pdf_name = self.ui.lineEdit_2.text()
            self.unmerge_cell(excel_name, "sheet1")
            self.worker = SearchPDFWorker("./化妆品.xlsx", pdf_name)
            self.worker.update_text.connect(self.update_text_edit)
            self.worker.update_text_browser.connect(self.update_text_browser)
            self.worker.finished.connect(self.on_search_finished)
            self.worker.start()

    def getText(self):
        text, ok = QInputDialog.getText(self, "密码", "输入密码")
        return text

    def update_text_edit(self, text):
        self.ui.textEdit.append(text)
        self.ui.textEdit.ensureCursorVisible()  # 确保光标可见，以自动滚动到底部

    def update_text_browser(self, text):
        self.ui.textBrowser.append(text)
        self.ui.textBrowser.ensureCursorVisible()  # 确保光标可见，以自动滚动到底部

    def on_search_finished(self):
        QMessageBox.information(self, '成功', '查找结束', QMessageBox.Ok)

    def unmerge_cell(self, excel_name, sheet_name):
        wb = openpyxl.load_workbook(excel_name)
        sheet = wb[sheet_name]
        merged_cells = list(sheet.merged_cells.ranges)
        for merged_cell in merged_cells.copy():
            merged_value = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
            sheet.unmerge_cells(str(merged_cell))
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                for column in range(merged_cell.min_col, merged_cell.max_col + 1):
                    cell = sheet.cell(row=row, column=column)
                    cell.value = merged_value
        wb.save('化妆品.xlsx')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    mywin = QmyWindow()
    mywin.show()
    sys.exit(app.exec_())
