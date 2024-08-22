import os
import pdfplumber
import re
import pandas as pd
# 第二步导入相应的模块
from PyPDF2 import PdfFileReader
import fitz

import os
import openpyxl
# pip install  pdfplumber


def  pdf_2_image(i,image_path,pdf_name):
    PDFdoc = fitz.open(pdf_name)  # 读取PDF文件
    page = PDFdoc[i]
    # 增强图片分辨率
    zoom_x = 3  # 水平方向
    zoom_y = 3  # 垂直方向
    mat = fitz.Matrix(zoom_x, zoom_y)
    pix = page.get_pixmap(matrix=mat)
    # 按原PDF名称新建文件夹并按顺序保存图片
    if not os.path.exists(image_path):  # 判断文件夹是否已存在
        os.makedirs(image_path)  # 不存在则新建，存在就跳过这行

    pix.save(image_path + "第"+str(i+1) + "页.png")  # 按PDF中的页面顺序命名并保存图片


def  write_txt(file_path,text):

    # 打开文件进行追加
    with open(file_path, 'a' ,encoding = 'utf-8') as file:

        file.write(text + '\n')  # 写入内容后跟一个换行符
def extract_text_info(filepath,pages):
    """
    提取PDF中的文字
    @param filepath:文件路径
    @return:
    """
    # print("正在提取第"+str(pages)+"页")
    with pdfplumber.open(filepath) as pdf:
        # 获取第2页数据
        page = pdf.pages[pages]

        text=page.extract_text()

        # print(text)
        # print(type(text))

        return text

        # for page in pdf.pages:
        #     print(page.extract_text())

        # page = pdf.pages[1]
        # print(page.extract_text())





# 第三步：定义相对应的函数
def get_num_pages(file_path):
    """
    获取文件总页码
    :param file_path: 文件路径
    :return:
    """
    reader = PdfFileReader(file_path)
    # # 不解密可能会报错：PyPDF2.utils.PdfReadError: File has not been decrypted
    # if reader.isEncrypted:
    #     reader.decrypt('')
    page_num = reader.getNumPages()
    return page_num

def match_pattern(pattern,text,i,pdf_name):

    matches = re.findall(pattern, text)

    if matches:

        # print(pattern)
        #print(f"找到了'{pattern}':"+str(i+1)+"页")
        pdf_2_image(i, "./img/",pdf_name)
        write_txt("./img/"+"第"+str(i+1)+"页"+".txt",pattern)
        return 1
    else:
        return 0
# 判断是否为nan值
def check_nan(value):
    return value != "nan"  # 判断是否为nan值，nan值会返回True
def search_pdf(excel_name,pdf_name):
    flag2list=[]
    count=get_num_pages(pdf_name)
    for i in range(count):

        text=extract_text_info(pdf_name,i)
        df = pd.read_excel(excel_name)
        # cont_text=text.replace(" ", "").strip()
        cont_text=text.strip()

        # print(df.head())

        for index, row in df.iterrows():

            # print(index)
            pa=str(row[2])
            # pa =pa.replace(" ", "").strip()
            pa = pa.strip()
            if check_nan(pa):
                # print(pa)
                flag2=match_pattern(pa,cont_text,i,pdf_name)
                if flag2==1:
                    print(i)
                    print(pa)
                    flag2list.append(flag2)
                    print("找到了=====>>>>>>>")
    os.remove("化妆品.xlsx")

    return flag2list
            # for j in range(row.size):
            #     if j < row.size - 1:
            #         j += 1
            #         pipei=row[j].replace(" ", "").strip()
            #         match_pattern(pipei,text,i,pdf_name)
def unmerge_cell(excel_name, sheet_name):
    # 打开工作簿并获取sheet
    wb = openpyxl.load_workbook(excel_name)
    sheet = wb[sheet_name]

    merged_cells = list(sheet.merged_cells.ranges)

    # 遍历每个合并单元格
    for merged_cell in merged_cells.copy():
        # 获取合并单元格的值
        merged_value = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col).value

        # 对'合并单元格'进行拆分
        sheet.unmerge_cells(str(merged_cell))

        # 将值分配给每个单元格
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            for column in range(merged_cell.min_col, merged_cell.max_col + 1):
                cell = sheet.cell(row=row, column=column)
                cell.value = merged_value

    # 保存工作簿
    wb.save('化妆品.xlsx')
# excel_name = 'INGREDIENT LIST.xlsx'
# sheet_name = 'sheet1'
# unmerge_cell(excel_name, sheet_name)
# search_pdf("./化妆品.xlsx","./Policy.pdf")



















